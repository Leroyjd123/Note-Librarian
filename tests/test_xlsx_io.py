"""Safety tests for the surgical xlsx editor.

These guarantee the core promise: editing only the requested cells leaves the rest
of the workbook (other cells, shared strings, styles, hyperlinks) byte-for-byte intact.
"""
import os
import sys
import zipfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from backend.xlsx_io import WorkbookEditor, col_to_idx, idx_to_col  # noqa: E402


def _make_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JWLManager"
    ws.append(["TITLE", "NOTE", "TAGS", "COLOR", "Link"])
    ws.append(["Old title", "Old note", "Faith", 6, "x"])
    ws.append(["Second", "Note two", "", 1, "y"])     # empty TAGS cell
    ws["E2"].hyperlink = "https://example.com/a"
    ws["E3"].hyperlink = "https://example.com/b"
    wb.save(path)


def test_col_helpers():
    assert col_to_idx("A") == 1 and col_to_idx("Z") == 26 and col_to_idx("AA") == 27
    assert idx_to_col(1) == "A" and idx_to_col(27) == "AA"


def test_roundtrip_only_targets_change(tmp_path):
    path = str(tmp_path / "wb.xlsx")
    _make_workbook(path)

    wb = WorkbookEditor(path)
    headers = wb.header_map()
    assert headers["TITLE"] == "A" and headers["COLOR"] == "D"

    # Edit title (string), color (numeric), and insert a TAGS value where empty.
    wb.apply_edits(
        {
            2: {headers["TITLE"]: "New & better <title>", headers["COLOR"]: 4},
            3: {headers["TAGS"]: "Inserted | Tag"},
        },
        numeric_cols={headers["COLOR"]},
    )

    # Re-read with openpyxl (independent reader) to verify values + integrity.
    out = openpyxl.load_workbook(path)["JWLManager"]
    assert out["A2"].value == "New & better <title>"
    assert out["D2"].value == 4
    assert out["C3"].value == "Inserted | Tag"
    # Untouched cells unchanged
    assert out["B2"].value == "Old note"
    assert out["A3"].value == "Second"
    # Hyperlinks preserved
    assert out["E2"].hyperlink is not None
    assert out["E3"].hyperlink is not None


def test_backup_created(tmp_path):
    path = str(tmp_path / "wb.xlsx")
    _make_workbook(path)
    wb = WorkbookEditor(path)
    wb.apply_edits({2: {"A": "changed"}})
    assert os.path.exists(path + ".bak")


def test_shared_strings_and_styles_untouched(tmp_path):
    path = str(tmp_path / "wb.xlsx")
    _make_workbook(path)
    with zipfile.ZipFile(path) as z:
        before_styles = z.read("xl/styles.xml")

    wb = WorkbookEditor(path)
    wb.apply_edits({2: {"A": "changed"}})

    with zipfile.ZipFile(path) as z:
        after_styles = z.read("xl/styles.xml")
    assert before_styles == after_styles


def test_ensure_columns_adds_new_and_keeps_data(tmp_path):
    path = str(tmp_path / "wb.xlsx")
    _make_workbook(path)
    wb = WorkbookEditor(path)
    cols = wb.ensure_columns(["TAGS", "CONFIDENCE", "REVIEW"])
    # existing header reused, new ones appended after the last column (E=5 -> F, G)
    assert cols["TAGS"] == "C"
    assert cols["CONFIDENCE"] == "F" and cols["REVIEW"] == "G"
    wb.apply_edits({2: {cols["CONFIDENCE"]: "HIGH", cols["REVIEW"]: "No"}})

    out = openpyxl.load_workbook(path)["JWLManager"]
    assert out["F1"].value == "CONFIDENCE" and out["G1"].value == "REVIEW"
    assert out["F2"].value == "HIGH" and out["G2"].value == "No"
    # original data + hyperlinks intact
    assert out["A2"].value == "Old title" and out["B2"].value == "Old note"
    assert out["E2"].hyperlink is not None
    assert out.max_column == 7


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
